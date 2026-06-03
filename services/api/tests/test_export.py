"""Phase 4 export gates. Deterministic and secret-free: each builds the workbook
from a fixture sheet payload and reads it back, with no database and no model.

Covered:
  fidelity   the title block, the sections in order, every value present, and a
             section-organized layout rather than a dump of the tidy store.
  styling    category-colored section headers, banded rows, currency and percent
             number formats, the merged title, and frozen panes.
  charts     a native chart of the right type for each chart-hinted section with
             enough points, and none for a non-chart section or too few points.
  comments   every value carries a comment with its source sentence and confidence.
  csv        the secondary flat format contains the visible values.

All values below are invented and labeled sample; none is real research.
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime

from openpyxl import load_workbook

from app.export import build_csv, build_xlsx, palette
from app.models import Cell, ColumnDef, SectionWithCells, Sheet, SheetPayload

FUNDING_SNIPPET = "SAMPLE: Acme raised 240 million rupees in its Series B."


def _cell(**over: object) -> Cell:
    base: dict[str, object] = {
        "id": "cell",
        "section_id": "section",
        "row_idx": 0,
        "col_key": None,
        "value_raw": None,
        "value_norm": None,
        "unit": None,
        "period": None,
        "source_snippet": "SAMPLE: supporting sentence for this value.",
        "char_start": None,
        "char_end": None,
        "confidence": None,
    }
    base.update(over)
    return Cell(**base)  # type: ignore[arg-type]


def _section(**over: object) -> SectionWithCells:
    base: dict[str, object] = {
        "id": "section",
        "sheet_id": "sheet",
        "key": "k",
        "label": "L",
        "kind": "scalar",
        "render_hint": "keyvalue",
        "category": None,
        "columns": None,
        "sort": 0,
        "confidence": None,
        "cells": [],
    }
    base.update(over)
    return SectionWithCells(**base)  # type: ignore[arg-type]


def _sample_payload() -> SheetPayload:
    sheet = Sheet(
        id="sheet",
        document_id="doc",
        title="Acme Robotics (sample)",
        status="done",
        field_count=0,
        cost_usd=0,
        created_at=datetime(2026, 1, 1),
    )
    sections = [
        _section(
            key="overview",
            label="Overview",
            kind="scalar",
            render_hint="keyvalue",
            category="Overview",
            cells=[
                _cell(
                    col_key="Funding",
                    value_raw="240000000",
                    value_norm="240000000",
                    unit="INR",
                    confidence=0.95,
                    source_snippet=FUNDING_SNIPPET,
                ),
                _cell(
                    col_key="Margin",
                    value_raw="45%",
                    value_norm="45",
                    unit="percent",
                    confidence=0.6,
                    row_idx=1,
                ),
            ],
        ),
        _section(
            key="metrics",
            label="Metrics",
            kind="table",
            render_hint="table",
            category="Financials",
            columns=[
                ColumnDef(key="metric", label="Metric"),
                ColumnDef(key="value", label="Value"),
            ],
            cells=[
                _cell(
                    row_idx=0,
                    col_key="metric",
                    value_raw="Revenue",
                    value_norm="Revenue",
                    confidence=0.8,
                ),
                _cell(row_idx=0, col_key="value", value_raw="10", value_norm="10", confidence=0.8),
                # Row 1 has no value cell: the export shows the neutral placeholder.
                _cell(
                    row_idx=1,
                    col_key="metric",
                    value_raw="EBITDA",
                    value_norm="EBITDA",
                    confidence=0.7,
                ),
                # Row 2 carries a cell whose column key was not declared: still shown.
                _cell(
                    row_idx=2, col_key="metric", value_raw="Note", value_norm="Note", confidence=0.7
                ),
                _cell(
                    row_idx=2,
                    col_key="remark",
                    value_raw="Strong demand (sample)",
                    value_norm="Strong demand (sample)",
                    confidence=0.6,
                ),
            ],
        ),
        _section(
            key="revenue_split",
            label="Revenue split",
            kind="table",
            render_hint="breakdown_pie",
            category="Financials",
            columns=[
                ColumnDef(key="region", label="Region"),
                ColumnDef(key="share", label="Share"),
            ],
            cells=[
                _cell(
                    row_idx=0, col_key="region", value_raw="EMEA", value_norm="EMEA", confidence=0.8
                ),
                _cell(
                    row_idx=0,
                    col_key="share",
                    value_raw="45%",
                    value_norm="45",
                    unit="percent",
                    confidence=0.8,
                ),
                _cell(
                    row_idx=1,
                    col_key="region",
                    value_raw="Americas",
                    value_norm="Americas",
                    confidence=0.8,
                ),
                _cell(
                    row_idx=1,
                    col_key="share",
                    value_raw="35%",
                    value_norm="35",
                    unit="percent",
                    confidence=0.8,
                ),
                _cell(
                    row_idx=2, col_key="region", value_raw="APAC", value_norm="APAC", confidence=0.8
                ),
                _cell(
                    row_idx=2,
                    col_key="share",
                    value_raw="20%",
                    value_norm="20",
                    unit="percent",
                    confidence=0.7,
                ),
            ],
        ),
        _section(
            key="arr_history",
            label="ARR history",
            kind="timeseries",
            render_hint="timeseries_bar",
            category="Financials",
            columns=[ColumnDef(key="arr", label="ARR")],
            cells=[
                _cell(
                    row_idx=0,
                    col_key="arr",
                    period="2021",
                    value_raw="$6M",
                    value_norm="6000000",
                    unit="USD",
                    confidence=0.7,
                ),
                _cell(
                    row_idx=1,
                    col_key="arr",
                    period="2022",
                    value_raw="$9M",
                    value_norm="9000000",
                    unit="USD",
                    confidence=0.7,
                ),
                _cell(
                    row_idx=2,
                    col_key="arr",
                    period="2023",
                    value_raw="$18M",
                    value_norm="18000000",
                    unit="USD",
                    confidence=0.8,
                ),
            ],
        ),
        # A chart-hinted section with only two points: too few for a sound chart,
        # so the export must show its table and no chart (the conservative rule).
        _section(
            key="sparse_series",
            label="Sparse series",
            kind="timeseries",
            render_hint="timeseries_bar",
            category="Financials",
            columns=[ColumnDef(key="arr", label="ARR")],
            cells=[
                _cell(
                    row_idx=0,
                    col_key="arr",
                    period="2022",
                    value_raw="$9M",
                    value_norm="9000000",
                    confidence=0.6,
                ),
                _cell(
                    row_idx=1,
                    col_key="arr",
                    period="2023",
                    value_raw="$18M",
                    value_norm="18000000",
                    confidence=0.6,
                ),
            ],
        ),
        _section(
            key="why_now",
            label="Why now",
            kind="longtext",
            render_hint="longtext",
            category="Thesis",
            cells=[
                _cell(
                    value_raw="Mid-market firms are moving to modern software (sample).",
                    value_norm="Mid-market firms are moving to modern software (sample).",
                    confidence=0.55,
                )
            ],
        ),
        _section(
            key="investors",
            label="Investors",
            kind="list",
            render_hint="chips",
            category="Capital",
            cells=[_cell(value_raw="Harbor Lane Ventures (sample)", confidence=0.7)],
        ),
    ]
    return SheetPayload(sheet=sheet, sections=sections)


def _build(payload: SheetPayload) -> bytes:
    return build_xlsx(
        payload,
        doc_name="Acme Robotics (sample).pdf",
        doc_type="company_profile",
        primary_topic="Acme Robotics, a robotics company (sample)",
    )


def _all_text_tokens(ws: object) -> set[str]:
    """Every cell value in the sheet, numbers rendered the way a value is."""
    found: set[str] = set()
    for row in ws.iter_rows():  # type: ignore[attr-defined]
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, (int, float)):
                found.add(str(int(value)) if float(value).is_integer() else str(value))
            else:
                found.add(str(value))
    return found


def _expected_value_token(cell: Cell) -> str:
    number = palette.numeric_value(cell.value_norm)
    if number is not None:
        return str(int(number)) if number.is_integer() else str(number)
    return palette.display_text(cell)


def test_export_fidelity() -> None:
    payload = _sample_payload()
    wb = load_workbook(io.BytesIO(_build(payload)))
    ws = wb.active

    # Title block: the subject and a classification subtitle.
    assert ws["A1"].value == "Acme Robotics (sample)"
    assert "Company profile" in str(ws["A2"].value)

    # Sections appear in the engine's order, each introduced by its header row.
    header_rows: list[int] = []
    for section in payload.sections:
        match = next(
            (c.row for r in ws.iter_rows() for c in r if c.value == section.label),
            None,
        )
        assert match is not None, f"section header missing: {section.label}"
        header_rows.append(match)
    assert header_rows == sorted(header_rows), "sections are not in engine order"

    # Every grounded value is present (numbers as written, text as written), so
    # the export does not lose data.
    found = _all_text_tokens(ws)
    for section in payload.sections:
        for cell in section.cells:
            token = _expected_value_token(cell)
            assert token in found, f"value missing from export: {token!r}"


def test_export_styling() -> None:
    payload = _sample_payload()
    wb = load_workbook(io.BytesIO(_build(payload)))
    ws = wb.active

    # A section header is filled with its category's accent, the same accent the
    # section has on screen (the palette is ported from the web client).
    accent = palette.category_accent("Overview").lstrip("#").upper()
    overview_fill = next(
        (c.fill.fgColor.rgb for r in ws.iter_rows() for c in r if c.value == "Overview"),
        None,
    )
    assert overview_fill is not None and str(overview_fill).upper().endswith(accent)

    # Banded table rows: at least one body cell carries the band fill.
    band = palette.BAND_BG.lstrip("#").upper()
    band_hits = sum(
        1
        for r in ws.iter_rows()
        for c in r
        if isinstance(c.fill.fgColor.rgb, str) and c.fill.fgColor.rgb.upper().endswith(band)
    )
    assert band_hits > 0, "no banded rows found"

    # Number formats: a rupee currency format and a percent format are applied.
    formats = {c.number_format for r in ws.iter_rows() for c in r if c.number_format}
    assert any("₹" in fmt or "#,##,##0" in fmt for fmt in formats), "no rupee format"
    assert any("%" in fmt for fmt in formats), "no percent format"

    # The merged title and the frozen header.
    assert any(rng.min_row == 1 and rng.min_col == 1 for rng in ws.merged_cells.ranges)
    assert ws.freeze_panes == "A3"


def test_export_native_charts() -> None:
    payload = _sample_payload()
    data = _build(payload)

    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        chart_parts = [
            name
            for name in archive.namelist()
            if name.startswith("xl/charts/chart") and name.endswith(".xml")
        ]
        kinds: list[str] = []
        for name in chart_parts:
            content = archive.read(name).decode("utf-8")
            if "<c:pieChart" in content:
                kinds.append("pie")
            elif "<c:barChart" in content:
                kinds.append("bar")
            elif "<c:lineChart" in content:
                kinds.append("line")

    # Exactly the two valid chart sections produced a chart, of the right type:
    # the breakdown is a pie and the three-point series is a bar. The two-point
    # series and the non-chart sections produced none.
    assert sorted(kinds) == ["bar", "pie"]
    assert len(chart_parts) == 2


def test_export_source_comments() -> None:
    payload = _sample_payload()
    wb = load_workbook(io.BytesIO(_build(payload)))
    ws = wb.active

    # The funding value cell carries a comment with its exact source sentence and
    # its confidence, so the grounding travels into the downloaded file.
    funding_comment = next(
        (c.comment.text for r in ws.iter_rows() for c in r if c.value == 240000000 and c.comment),
        None,
    )
    assert funding_comment is not None
    assert FUNDING_SNIPPET in funding_comment
    assert "Confidence" in funding_comment


def test_export_csv_secondary() -> None:
    payload = _sample_payload()
    text = build_csv(payload).decode("utf-8")

    assert text.splitlines()[0] == "Section,Field,Period,Value,Unit"
    for token in ("Overview", "Revenue", "Harbor Lane Ventures (sample)", "45%"):
        assert token in text, f"csv missing visible value: {token!r}"
